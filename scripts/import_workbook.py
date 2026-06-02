"""
Import the pump configurator workbook into tuskdb.sqlite3.

Sources (faithful to Option A schema):
  * "Config Info"  -> one catalog PumpConfig per row (PumpInfo + Seal + Motor)
  * "Price Lists"  -> PriceList (base prices) + OptionPrice (per-option add-ons)

Run:
    .venv/Scripts/python.exe scripts/import_workbook.py

The script is idempotent:
  * PriceList / OptionPrice use update_or_create on their natural keys.
  * Catalog PumpConfigs are matched by `name` (the Configuration ID); existing
    ones are skipped so re-running won't create duplicates.
"""

from __future__ import annotations

import asyncio
import os
import sys
from decimal import Decimal

import openpyxl
from tortoise import Tortoise

# Make `src` importable when run from the project root.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.database import TORTOISE_ORM  # noqa: E402
from src.models import (  # noqa: E402
    Motor,
    OptionPrice,
    PriceList,
    PumpConfig,
    PumpInfo,
    Seal,
)

WORKBOOK = "pump_configurator_updated_with_prices.xlsx"


def _s(value) -> str | None:
    """Normalise a cell value to a trimmed string (or None when empty)."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _dec(value) -> Decimal:
    if value is None or _s(value) is None:
        return Decimal("0")
    return Decimal(str(value))


async def import_config_info(ws) -> tuple[int, int]:
    """Create a catalog PumpConfig (+ PumpInfo/Seal/Motor) per Config Info row."""
    created = skipped = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        config_id = _s(row[0])
        if not config_id:
            continue  # blank line

        if await PumpConfig.get_or_none(name=config_id, is_catalog=True):
            skipped += 1
            continue

        pump_info = await PumpInfo.create(
            series=_s(row[1]) or "",
            size=_s(row[2]) or "",
            pump_material=_s(row[3]) or "",
            shaft_configuration=_s(row[4]) or "",
            casing_metal=_s(row[5]) or "",
            casing_drain=_s(row[6]) or "",
            casing_tap=_s(row[7]) or "",
            casing_gasket=_s(row[8]) or "Grafoil",
            flange_configuration=_s(row[9]) or "",
            casing_wear_ring=_s(row[10]) or "",
        )

        seal = await Seal.create(
            seal_configuration=_s(row[11]) or "",
            seal_type=_s(row[12]) or "",
            inboard_rotating_face=_s(row[13]) or "",
            inboard_stationary_face=_s(row[14]) or "",
            inboard_elastomer=_s(row[15]) or "",
        )

        motor = await Motor.create(
            power_hp=_s(row[16]) or "",
            speed=_s(row[17]) or "",
            voltage=_s(row[18]) or "",
        )

        await PumpConfig.create(
            name=config_id,
            notes=_s(row[19]),
            is_catalog=True,
            list_price=None,         # Config Info carries no per-config price
            pump_info=pump_info,
            seal=seal,
            motor=motor,
        )
        created += 1

    return created, skipped


async def import_price_lists(ws) -> tuple[int, int]:
    """Load the base-price block (A-D) and option-price block (F-I)."""
    base_count = option_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Left block: Key | Product Family | Size | Base Price  (cols 0-3)
        family, size, base_price = _s(row[1]), _s(row[2]), row[3]
        if family and size:
            await PriceList.update_or_create(
                product_family=family,
                size=size,
                defaults={"base_price": _dec(base_price)},
            )
            base_count += 1

        # Middle block: Key | Field | Option | Option Price  (cols 5-8)
        field, option, option_price = _s(row[6]), _s(row[7]), row[8]
        if field and option:
            await OptionPrice.update_or_create(
                field=field,
                option=option,
                defaults={"option_price": _dec(option_price)},
            )
            option_count += 1

    return base_count, option_count


async def main() -> None:
    if not os.path.exists(WORKBOOK):
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    await Tortoise.init(config=TORTOISE_ORM)
    try:
        created, skipped = await import_config_info(wb["Config Info"])
        bases, options = await import_price_lists(wb["Price Lists"])
    finally:
        await Tortoise.close_connections()

    print("Import complete:")
    print(f"  catalog PumpConfigs : {created} created, {skipped} skipped (already present)")
    print(f"  PriceList rows      : {bases}")
    print(f"  OptionPrice rows    : {options}")


if __name__ == "__main__":
    asyncio.run(main())
